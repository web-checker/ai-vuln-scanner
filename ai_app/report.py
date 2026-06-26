"""
최종 보고서 생성 (CSV / Excel).

Excel(.xlsx) 은 외부 양식을 그대로 따른 5개 시트로 구성한다:
  0. 표지            - 문서 메타정보 + 제목(기본값, 엑셀에서 수정 가능)
  1. 진단 대상       - 서버(Hostname/IP/버전/용도) 목록
  2-1. 요약결과(그래프) - 도메인 평균점수 / 양호·취약·N/A / 영역별 취약수 (+ 차트)
  2-2. 요약 진단결과 - 항목별 결과표 + 영역별 점수 + 보안 적용율
  3-1. 진단 결과     - 항목별 상세(진단기준/결과/비고)

점수 규칙: 영역점수 = 양호 / (양호 + 취약)  (N/A 제외). 보안 적용율도 동일.
모두 순수 파이썬(LLM 호출 없음) → 토큰 0.
"""
from __future__ import annotations

import io
from datetime import date

import pandas as pd

from . import config
from .preprocess import restore_multiline

# ── 표지/메타 기본값 (CSV에 없는 값 — 엑셀에서 직접 수정 가능) ──
REPORT_META = {
    "문서번호": "CHECKBANG-VA-2026-001",
    "작성자": "취약점진단팀",
    "보안등급": "Confidential",
    "Ver": "ver 1.0",
    "제목": 'Check방 취약점 진단',
    "부제": "취약점 진단 상세결과",
}
# 진단대상 시트에서 CSV에 없는 칸의 기본값(용도는 수기 기입용으로 공란)
TARGET_DEFAULTS = {"버전정보": "-", "용도": "", "비고": "-"}


def _label_reason(result: str, reason: str) -> str:
    """판단근거 본문만 반환. 결과(양호/취약)는 '결과' 열에 이미 표기되므로 머리말은 붙이지 않는다."""
    return (reason or "").strip()


def _report_row(row, *, result: str, reason: str, remediation: str, show_remed: bool) -> dict:
    """원본 CSV / Run CSV 공통 메타 컬럼 + 파생 판정값으로 보고서 행(REPORT_COLUMNS) 생성.

    분류·중요도·항목·판단기준·진단대상(IP)은 두 소스가 같은 컬럼명을 쓰므로 그대로 읽고,
    result/reason/remediation/show_remed 만 호출측이 소스별로 파생해 넘긴다.
    조치방법은 show_remed(스크립트/AI 중 하나라도 취약)일 때만 표기.
    """
    return {
        "항목코드": row.get("항목코드", ""),
        "분류": row.get("분류", ""),
        "중요도": row.get("중요도", ""),
        "항목": row.get("항목", ""),
        "판단기준": _crit_lines(row.get("판단기준", "")),
        "결과": result,
        "판단근거": _label_reason(result, reason),
        "조치방법": remediation if show_remed else "",
        "진단대상": row.get("진단대상", ""),
        "진단대상IP": row.get("진단대상IP", ""),
    }


def build_report_df(df: pd.DataFrame, decisions: dict[str, dict]) -> pd.DataFrame:
    """원본 CSV DataFrame + 확정 판정(decisions)을 보고서 DataFrame으로 변환."""
    rows = []
    for _, row in df.iterrows():
        dec = decisions.get(row.get("항목코드", ""), {})
        result = config.final_result(dec.get("result", ""), row.get("결과", ""))
        reason = dec.get("reason") or restore_multiline(row.get("점검내용", ""))
        rows.append(_report_row(
            row, result=result, reason=reason,
            remediation=dec.get("remediation", ""),
            show_remed=dec.get("vuln_any", result == config.R_VULN)))
    return pd.DataFrame(rows, columns=config.REPORT_COLUMNS)


def to_csv_bytes(report_df: pd.DataFrame) -> bytes:
    """보고서 DataFrame → CSV 바이트(UTF-8 BOM, 전체 인용)."""
    import csv

    csv_text = report_df.to_csv(index=False, quoting=csv.QUOTE_ALL, lineterminator="\n")
    return b"\xef\xbb\xbf" + csv_text.encode("utf-8")


# ════════════════════════════════════════════════════════════════
#  Excel 보고서 (5시트 양식)
# ════════════════════════════════════════════════════════════════
SUMMARY_SHEET = "2-2. 요약 진단결과"   # 수식 참조용 시트 이름


def build_xlsx_from_report_df(rdf: pd.DataFrame) -> bytes:
    """보고서 DataFrame(REPORT_COLUMNS) → 5시트 .xlsx 바이트."""
    wb = _build_workbook(rdf)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_xlsx_report(df: pd.DataFrame, decisions: dict[str, dict]) -> bytes:
    """원본 df + 확정 판정 → 5시트 .xlsx 바이트(수식 자동계산 포함)."""
    return build_xlsx_from_report_df(build_report_df(df, decisions))


def build_compare_csv(compare_df: pd.DataFrame) -> bytes:
    """비교 결과 DataFrame(COMPARE_COLUMNS)을 CSV 바이트로(UTF-8 BOM + 전체 인용)."""
    return to_csv_bytes(compare_df)


# ════════════════════════════════════════════════════════════════
#  HTML 보고서 (자체 완결형 — 인라인 CSS, 오프라인 열람·인쇄 가능)
# ════════════════════════════════════════════════════════════════
def _esc(text) -> str:
    """HTML 이스케이프 + 줄바꿈을 <br>로."""
    import html as _html
    return _html.escape(str(text or "")).replace("\n", "<br>")


def _guide_remediation(code: str) -> str:
    """가이드 PDF의 '조치 방법' 절 자동추출(저장값 없을 때 폴백). 실패 시 ''."""
    try:
        from . import guide_index
        path = config.guide_pdf_for_code(code)
        if path is None:
            return ""
        txt = guide_index.remediation_section(str(path), code)
        return f"[가이드 권고]\n{txt}" if txt else ""
    except Exception:  # noqa: BLE001  (PDF 파싱 실패 무시)
        return ""


def build_report_df_from_run(run_df: pd.DataFrame) -> pd.DataFrame:
    """저장된 Run CSV(RUN_COLUMNS) → 보고서 DataFrame(REPORT_COLUMNS).

    최종결과=확정값 우선·없으면 스크립트결과, 근거=확정근거 우선·없으면 AI근거.
    스크립트/AI 중 하나라도 취약이면 조치방법 표기. 저장된 조치방법이 비었으면
    주통기 가이드 PDF의 '조치 방법' 절을 자동 추출해 채운다.
    """
    rows = []
    for _, r in run_df.iterrows():
        script, ai = r.get("스크립트결과", ""), r.get("AI결과", "")
        result = config.final_result(r.get("확정결과", ""), script)
        reason = r.get("확정근거", "") or r.get("AI근거", "")
        vuln = config.R_VULN in (script, ai)
        remediation = r.get("조치방법", "")
        if vuln and not str(remediation).strip():
            remediation = _guide_remediation(str(r.get("항목코드", "")))
        rows.append(_report_row(
            r, result=result, reason=reason,
            remediation=remediation, show_remed=vuln))
    return pd.DataFrame(rows, columns=config.REPORT_COLUMNS)


# ── 최종 보고서(최초진단 base ↔ 이행점검 target 병합) ──────────────
def build_final_report_rows(base_df: pd.DataFrame, target_df: pd.DataFrame) -> list[dict]:
    """두 Run을 항목코드로 병합 → 최종 보고서 표 행(UI용).

    최초결과/최초근거 = base(최초진단), 최종결과/최종근거 = target(이행점검).
    """
    brdf = build_report_df_from_run(base_df)
    trdf = build_report_df_from_run(target_df)
    tmap = {str(r["항목코드"]): r for _, r in trdf.iterrows()}
    out = []
    for _, b in brdf.iterrows():
        code = str(b["항목코드"])
        t = tmap.get(code)
        out.append({
            "code": code, "group": b["분류"], "severity": b["중요도"], "name": b["항목"],
            "criteria": b["판단기준"],
            "firstResult": b["결과"], "firstReason": b["판단근거"],
            "finalResult": (t["결과"] if t is not None else ""),
            "finalReason": (t["판단근거"] if t is not None else ""),
            "target": b["진단대상"], "ip": b["진단대상IP"],
        })
    return out


def build_final_xlsx(base_df: pd.DataFrame, target_df: pd.DataFrame) -> bytes:
    """최종 보고서 5시트 xlsx.

    요약/그래프(0·1·2-1·2-2)는 '최종'(이행점검=target) 결과 기준으로 집계되고,
    3-1 진단 결과 시트만 최초(base)·최종(target) 결과·근거를 8컬럼으로 함께 보인다.
    """
    brdf = build_report_df_from_run(base_df)      # 최초진단
    trdf = build_report_df_from_run(target_df)    # 이행점검(최종) → 요약/그래프 기준
    base_map = {str(r["항목코드"]): {"result": r["결과"], "reason": r["판단근거"]}
                for _, r in brdf.iterrows()}
    wb = _build_workbook(trdf, compare_map=base_map)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_html_report(df: pd.DataFrame, decisions: dict[str, dict],
                      meta: dict | None = None) -> str:
    """원본 df + 확정 판정 → 자체 완결형 HTML 보고서 문자열(엑셀과 표 내용 일치)."""
    return _render_html_report(build_report_df(df, decisions), meta)


def build_html_report_compare(cmp: dict, meta: dict | None = None) -> str:
    """비교 결과(store.compare 반환 dict) → 자체 완결형 HTML 비교 보고서."""
    m = meta or {}
    s = cmp.get("summary", {})
    base = cmp.get("base") or {}
    target = cmp.get("target") or {}
    fix = s.get("fixRate")
    meta_rows = [
        ("진단 대상", m.get("대상", "") or "-"),
        ("기준 파일", f"{base.get('종류', '')} · {base.get('일시', '')} · {base.get('원본파일명', '')}"),
        ("비교 파일", f"{target.get('종류', '')} · {target.get('일시', '')} · {target.get('원본파일명', '')}"),
        ("생성 일시", m.get("일시", str(date.today()))),
    ]
    meta_html = "".join(
        f"<tr><th>{_esc(k)}</th><td>{_esc(v)}</td></tr>" for k, v in meta_rows)

    kpis = [
        ("조치 완료", s.get("improved", 0), "취약 → 양호", "k-pass"),
        ("미조치", s.get("unfixed", 0), "취약 → 취약", "k-vuln"),
        ("신규 취약", s.get("worsened", 0), "양호 → 취약", "k-vuln"),
        ("조치율", "-" if fix is None else f"{fix}%", f"기준 취약 {s.get('baseVuln', 0)}건", "k-total"),
    ]
    kpi_html = "".join(
        f'<div class="kcard {cls}"><div class="knum">{v}</div>'
        f'<div class="klabel">{_esc(label)}</div><div class="ksub">{_esc(sub)}</div></div>'
        for label, v, sub, cls in kpis)

    st_cls = {config.C_IMPROVED: "pass", config.C_KEPT: "pass", config.C_UNFIXED: "vuln",
              config.C_WORSENED: "warn", config.C_NA: "na"}
    rcls = {config.R_PASS: "pass", config.R_VULN: "vuln", config.R_NA: "na"}
    body_rows = []
    for r in cmp.get("rows", []):
        b = str(r.get("최초결과") or "")
        t = str(r.get("이행결과") or "")
        stt = str(r.get("상태") or "")
        body_rows.append(
            "<tr>"
            f'<td class="code">{_esc(r.get("항목코드"))}</td>'
            f'<td>{_esc(r.get("분류"))}</td>'
            f'<td class="nm">{_esc(r.get("항목"))}</td>'
            f'<td class="c"><span class="sev s-{_esc(r.get("중요도"))}">{_esc(r.get("중요도"))}</span></td>'
            f'<td class="c"><span class="pill {rcls.get(b, "na")}">{_esc(b or "-")}</span></td>'
            f'<td class="c"><span class="pill {rcls.get(t, "na")}">{_esc(t or "-")}</span></td>'
            f'<td class="c"><span class="pill {st_cls.get(stt, "na")}">{_esc(stt)}</span></td>'
            "</tr>")

    return _HTML_COMPARE_TEMPLATE.format(
        title="진단 결과 비교 보고서", meta_html=meta_html, kpi_html=kpi_html,
        body_rows="".join(body_rows))


def build_html_report_from_run(run_df: pd.DataFrame, meta: dict | None = None) -> str:
    """저장된 Run CSV → 자체 완결형 HTML 보고서 문자열."""
    return _render_html_report(build_report_df_from_run(run_df), meta)


def _render_html_report(rdf: pd.DataFrame, meta: dict | None = None) -> str:
    """보고서 DataFrame(REPORT_COLUMNS)을 자체 완결형 HTML 문자열로 렌더링.

    meta: {"종류","일시","원본파일명"} 등 상단 표기 정보(선택).
    """
    counts = _counts(rdf)
    targets = _targets(rdf)
    m = meta or {}
    has_base = (counts["pass"] + counts["vuln"]) > 0
    applied_pct = f"{counts['applied'] * 100:.1f}%" if has_base else "-"
    applied_w = round(counts["applied"] * 100, 1) if has_base else 0

    target_line = ", ".join(
        f"{t['hostname']} ({t['ip']})" for t in targets) or "-"
    meta_rows = [
        ("진단 대상", target_line),
        ("진단 종류", m.get("종류", "")),
        ("진단 일시", m.get("일시", str(date.today()))),
        ("원본 파일", m.get("원본파일명", "")),
    ]
    meta_html = "".join(
        f"<tr><th>{_esc(k)}</th><td>{_esc(v)}</td></tr>" for k, v in meta_rows if v != "")

    # 양호/취약/N·A 세그먼트 — 개수에 비례해 폭이 달라져 대략적 비율을 시각화(0건은 생략)
    segs = [("양호", counts["pass"], "s-pass"), ("취약", counts["vuln"], "s-vuln"),
            ("N/A", counts["na"], "s-na")]
    seg_html = "".join(
        f'<div class="seg {cls}" style="flex-grow:{n}">'
        f'<span class="segn">{n}</span><span class="segl">{_esc(label)}</span></div>'
        for label, n, cls in segs if n > 0)

    body_rows = []
    for _, r in rdf.iterrows():
        res = str(r.get("결과") or "")
        cls = {config.R_PASS: "pass", config.R_VULN: "vuln", config.R_NA: "na"}.get(res, "na")
        body_rows.append(
            "<tr>"
            f'<td class="code">{_esc(r.get("항목코드"))}</td>'
            f'<td>{_esc(r.get("분류"))}</td>'
            f'<td class="c"><span class="sev s-{_esc(r.get("중요도"))}">{_esc(r.get("중요도"))}</span></td>'
            f'<td class="nm">{_esc(r.get("항목"))}</td>'
            f'<td class="wrap">{_esc(r.get("판단기준"))}</td>'
            f'<td class="c"><span class="pill {cls}">{_esc(res or "N/A")}</span></td>'
            f'<td class="wrap">{_esc(r.get("판단근거"))}</td>'
            f'<td class="wrap">{_esc(r.get("조치방법"))}</td>'
            f'<td class="c">{_esc(r.get("진단대상"))}</td>'
            f'<td class="c">{_esc(r.get("진단대상IP"))}</td>'
            "</tr>")

    return _HTML_TEMPLATE.format(
        title=_esc(REPORT_META["제목"]), subtitle=_esc(REPORT_META["부제"]),
        meta_html=meta_html, total=counts["total"], seg_html=seg_html,
        applied_pct=_esc(applied_pct), applied_w=applied_w,
        body_rows="".join(body_rows))


_HTML_TEMPLATE = """<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title} - {subtitle}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 28px 24px 60px; background: #f4f6fb; color: #1f2738;
    font-family: "Malgun Gothic","맑은 고딕",-apple-system,Segoe UI,Roboto,sans-serif; font-size: 13px; }}
  .doc {{ max-width: 1400px; margin: 0 auto; }}
  h1 {{ font-size: 24px; margin: 0 0 4px; }}
  .sub {{ color: #5b677e; margin: 0 0 18px; font-weight: 600; }}
  .meta {{ flex: 1 1 0; min-width: 0; align-self: stretch; border-collapse: collapse; margin: 0; background: #fff;
    border: 1px solid #cdd7e8; border-radius: 12px; overflow: hidden; }}
  .meta th, .meta td {{ padding: 8px 12px; text-align: left; border-bottom: 1px solid #e3e9f3; }}
  .meta th {{ background: #d6e2f5; color: #1c3f74; width: 84px; font-weight: 700; }}
  .meta td {{ font-weight: 500; background: #f3f7fd; }}
  .meta tr:last-child th, .meta tr:last-child td {{ border-bottom: 0; }}
  /* 요약: 메타표 + (총항목 + 양호/취약/N·A 세그먼트) + 보안 적용율 — 한 줄에 1/3씩 */
  .summary {{ display: flex; gap: 14px; align-items: flex-start; margin: 0 0 22px; flex-wrap: wrap; }}
  .counts {{ flex: 1 1 0; min-width: 220px; background: #fff; border: 1px solid #cdd7e8;
    border-radius: 12px; padding: 14px 16px; }}
  .totbar {{ display: flex; align-items: center; justify-content: center; height: 26px; padding: 0 12px;
    border-radius: 7px; background: #1f4e78; color: #fff; font-weight: 800; font-size: 13px; margin-bottom: 8px; }}
  .segbar {{ display: flex; gap: 3px; height: 46px; }}
  .seg {{ display: flex; flex-direction: column; align-items: center; justify-content: center;
    color: #fff; min-width: 50px; border-radius: 7px; padding: 0 6px; }}
  .seg .segn {{ font-size: 18px; font-weight: 800; line-height: 1.05; }}
  .seg .segl {{ font-size: 11px; font-weight: 700; opacity: .95; }}
  .seg.s-pass {{ background: #2f9e54; }}
  .seg.s-vuln {{ background: #e23b3b; }}
  .seg.s-na {{ background: #9aa3b2; }}
  .appliedmini {{ flex: 1 1 0; min-width: 220px; align-self: stretch; background: #fff; border: 1px solid #cdd7e8;
    border-radius: 12px; padding: 14px 16px; display: flex; flex-direction: column; justify-content: center; }}
  .appliedmini .amlabel {{ color: #5b677e; font-weight: 700; font-size: 12px; }}
  .appliedmini .ampct {{ font-size: 22px; font-weight: 800; color: #1f4e78; margin: 2px 0 8px; }}
  .amtrack {{ height: 10px; background: #eef1f6; border-radius: 999px; overflow: hidden; }}
  .amfill {{ height: 100%; border-radius: 999px; background: linear-gradient(90deg, #2f9e54, #2563eb); }}
  table.report {{ width: 100%; border-collapse: collapse; background: #fff;
    border: 1px solid #d8deea; border-radius: 10px; overflow: hidden; table-layout: fixed; }}
  .report th, .report td {{ padding: 9px 10px; border-bottom: 1px solid #eef1f6;
    border-right: 1px solid #f1f3f8; vertical-align: middle; text-align: center; }}
  .report thead th {{ position: sticky; top: 0; background: #1f4e78; color: #fff;
    font-weight: 700; text-align: center; white-space: nowrap; }}
  .report tbody tr:nth-child(even) {{ background: #fafbfe; }}
  .report .c {{ text-align: center; }}
  .report .code {{ font-family: ui-monospace,Consolas,monospace; white-space: nowrap; }}
  .report .nm {{ font-weight: 600; }}
  .wrap {{ white-space: pre-wrap; word-break: break-word; line-height: 1.5; }}
  /* 열 너비: 긴 텍스트 열(판단기준/판단근거/조치방법)에 넉넉히 배분 */
  col.c-code {{ width: 78px; }} col.c-grp {{ width: 90px; }} col.c-sev {{ width: 56px; }}
  col.c-nm {{ width: 150px; }} col.c-crit {{ width: 16%; }} col.c-res {{ width: 72px; }}
  col.c-reason {{ width: 24%; }} col.c-remed {{ width: 22%; }}
  col.c-tgt {{ width: 90px; }} col.c-ip {{ width: 110px; }}
  .pill {{ display: inline-block; padding: 2px 10px; border-radius: 999px; font-weight: 700; font-size: 12px; white-space: nowrap; }}
  .pill.pass {{ background: #e2efda; color: #2f7d32; }}
  .pill.vuln {{ background: #fce4e4; color: #c0392b; }}
  .pill.na {{ background: #ededed; color: #555; }}
  .sev {{ display: inline-block; padding: 1px 8px; border-radius: 6px; font-size: 12px; font-weight: 700; }}
  .sev.s-상 {{ background: #fde2e1; color: #c0392b; }}
  .sev.s-중 {{ background: #fef3d6; color: #b8860b; }}
  .sev.s-하 {{ background: #e3ecfb; color: #2563eb; }}
  @media print {{
    body {{ background: #fff; padding: 0; font-size: 11px; }}
    .report thead th {{ position: static; }}
    table.report, .meta {{ border-radius: 0; }}
  }}
</style></head>
<body><div class="doc">
  <h1>{title}</h1>
  <p class="sub">{subtitle}</p>
  <div class="summary">
    <table class="meta">{meta_html}</table>
    <div class="counts">
      <div class="totbar">총 {total} 항목</div>
      <div class="segbar">{seg_html}</div>
    </div>
    <div class="appliedmini">
      <div class="amlabel">보안 적용율 (양호 / (양호 + 취약))</div>
      <div class="ampct">{applied_pct}</div>
      <div class="amtrack"><div class="amfill" style="width:{applied_w}%"></div></div>
    </div>
  </div>
  <table class="report">
    <colgroup>
      <col class="c-code"><col class="c-grp"><col class="c-sev"><col class="c-nm">
      <col class="c-crit"><col class="c-res"><col class="c-reason"><col class="c-remed">
      <col class="c-tgt"><col class="c-ip">
    </colgroup>
    <thead><tr>
      <th>항목코드</th><th>분류</th><th>중요도</th><th>항목</th><th>판단 기준</th>
      <th>결과</th><th>판단 근거</th><th>조치 방법</th><th>진단 대상</th><th>진단 대상 IP</th>
    </tr></thead>
    <tbody>{body_rows}</tbody>
  </table>
</div></body></html>"""


_HTML_COMPARE_TEMPLATE = """<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 28px 24px 60px; background: #f4f6fb; color: #1f2738;
    font-family: "Malgun Gothic","맑은 고딕",-apple-system,Segoe UI,Roboto,sans-serif; font-size: 13px; }}
  .doc {{ max-width: 1100px; margin: 0 auto; }}
  h1 {{ font-size: 24px; margin: 0 0 14px; }}
  .meta {{ border-collapse: collapse; margin: 0 0 18px; background: #fff; width: 100%;
    border: 1px solid #d8deea; border-radius: 10px; overflow: hidden; }}
  .meta th, .meta td {{ padding: 9px 14px; text-align: left; border-bottom: 1px solid #eef1f6; }}
  .meta th {{ background: #d6e2f5; color: #1c3f74; width: 120px; font-weight: 700; }}
  .meta td {{ font-weight: 500; background: #f3f7fd; }}
  .meta tr:last-child th, .meta tr:last-child td {{ border-bottom: 0; }}
  .kpis {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 0 0 18px; }}
  .kcard {{ flex: 1; min-width: 150px; background: #fff; border: 1px solid #d8deea;
    border-top: 4px solid #c9d2e3; border-radius: 12px; padding: 14px 16px; }}
  .kcard .knum {{ font-size: 28px; font-weight: 800; line-height: 1.1; }}
  .kcard .klabel {{ margin-top: 4px; color: #3a455c; font-weight: 700; font-size: 13px; }}
  .kcard .ksub {{ color: #7e8aa0; font-weight: 600; font-size: 11px; }}
  .kcard.k-total {{ border-top-color: #2563eb; }} .kcard.k-total .knum {{ color: #1f4e78; }}
  .kcard.k-vuln {{ border-top-color: #e23b3b; background: #fff6f6; }} .kcard.k-vuln .knum {{ color: #c0392b; }}
  .kcard.k-pass {{ border-top-color: #2f9e54; background: #f5fbf6; }} .kcard.k-pass .knum {{ color: #2f7d32; }}
  table.report {{ width: 100%; border-collapse: collapse; background: #fff;
    border: 1px solid #d8deea; border-radius: 10px; overflow: hidden; }}
  .report th, .report td {{ padding: 9px 10px; border-bottom: 1px solid #eef1f6;
    border-right: 1px solid #f1f3f8; vertical-align: top; }}
  .report thead th {{ background: #1f4e78; color: #fff; font-weight: 700; text-align: center; }}
  .report tbody tr:nth-child(even) {{ background: #fafbfe; }}
  .report .c {{ text-align: center; }}
  .report .code {{ font-family: ui-monospace,Consolas,monospace; white-space: nowrap; }}
  .report .nm {{ font-weight: 600; }}
  .pill {{ display: inline-block; padding: 2px 10px; border-radius: 999px; font-weight: 700; font-size: 12px; white-space: nowrap; }}
  .pill.pass {{ background: #e2efda; color: #2f7d32; }}
  .pill.vuln {{ background: #fce4e4; color: #c0392b; }}
  .pill.warn {{ background: #fef3d6; color: #b8860b; }}
  .pill.na {{ background: #ededed; color: #555; }}
  .sev {{ display: inline-block; padding: 1px 8px; border-radius: 6px; font-size: 12px; font-weight: 700; }}
  .sev.s-상 {{ background: #fde2e1; color: #c0392b; }}
  .sev.s-중 {{ background: #fef3d6; color: #b8860b; }}
  .sev.s-하 {{ background: #e3ecfb; color: #2563eb; }}
  @media print {{ body {{ background: #fff; padding: 0; }} }}
</style></head>
<body><div class="doc">
  <h1>{title}</h1>
  <table class="meta">{meta_html}</table>
  <div class="kpis">{kpi_html}</div>
  <table class="report">
    <thead><tr>
      <th>항목코드</th><th>분류</th><th>항목</th><th>중요도</th>
      <th>기준 결과</th><th>대상 결과</th><th>상태</th>
    </tr></thead>
    <tbody>{body_rows}</tbody>
  </table>
</div></body></html>"""


# ── 데이터 가공 헬퍼 ────────────────────────────────────────────
def _grouped(rdf: pd.DataFrame):
    """[(분류, [row, ...]), ...] — 등장 순서 유지."""
    groups: dict[str, list] = {}
    order: list[str] = []
    for _, r in rdf.iterrows():
        d = (str(r.get("분류") or "").strip()) or "기타"
        if d not in groups:
            groups[d] = []
            order.append(d)
        groups[d].append(r)
    return [(d, groups[d]) for d in order]


def _counts(rdf: pd.DataFrame) -> dict:
    # '결과' 컬럼이 없거나 빈 df 여도 KeyError 없이 0 집계(빈 보고서·임의 df 방어).
    res = [str(x) for x in rdf["결과"]] if "결과" in rdf.columns else []
    p = res.count(config.R_PASS)
    v = res.count(config.R_VULN)
    n = res.count(config.R_NA)
    return {"pass": p, "vuln": v, "na": n, "total": len(res),
            "applied": (p / (p + v) if (p + v) else 0)}


def _targets(rdf: pd.DataFrame) -> list[dict]:
    seen, out = [], []
    for _, r in rdf.iterrows():
        key = (str(r.get("진단대상", "") or ""), str(r.get("진단대상IP", "") or ""))
        if (key[0] or key[1]) and key not in seen:
            seen.append(key)
            out.append({"hostname": key[0] or "-", "ip": key[1] or "-"})
    return out or [{"hostname": "-", "ip": "-"}]


def _ranges(groups, data_start=6):
    """각 분류의 데이터 행 범위 [(분류, rows, r1, r2), ...] 계산."""
    out, r = [], data_start
    for d, rows in groups:
        out.append((d, rows, r, r + len(rows) - 1))
        r += len(rows)
    return out, data_start, r - 1, r   # ranges, start, end, footer_row


def _crit_lines(text) -> str:
    """판단기준의 '|' 를 줄바꿈으로 (빈 줄 없이)."""
    parts = [p.strip() for p in str(text or "").split("|")]
    return "\n".join(p for p in parts if p)


# ── 서식 상수/헬퍼 ──────────────────────────────────────────────
# Excel 워크북 빌더 — 다중시트/차트 보고서. build_xlsx_from_report_df → _build_workbook 가 사용.
def _xlsx_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="BFBFBF")
    return {
        "hdr_fill": PatternFill("solid", fgColor="1F4E78"),
        "hdr_font": Font(bold=True, color="FFFFFF"),
        "pass": PatternFill("solid", fgColor="E2EFDA"),
        "vuln": PatternFill("solid", fgColor="FCE4E4"),
        "na": PatternFill("solid", fgColor="EDEDED"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "leftc": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "topleft": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "Font": Font, "Alignment": Alignment,
    }
def _auto_width(ws, bounds=(5, 55), per_col=None):
    """열별 내용 길이에 맞춰 너비 자동 설정(병합·수식 셀 제외, 한글 2배)."""
    per_col = per_col or {}
    merged = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged.add((r, c))
    measured: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in merged:
                continue
            s = str(cell.value)
            if s.startswith("="):
                s = "000.0%"   # 수식 → 숫자 표시폭으로 가정
            ln = max((sum(2 if ord(ch) > 0x1100 else 1 for ch in line)
                      for line in s.split("\n")), default=0)
            col = cell.column_letter
            measured[col] = max(measured.get(col, 0), ln)
    for col, w in measured.items():
        lo, hi = per_col.get(col, bounds)
        ws.column_dimensions[col].width = min(max(w + 2, lo), hi)
def _result_fill(S, v):
    return {config.R_PASS: S["pass"], config.R_VULN: S["vuln"], config.R_NA: S["na"]}.get(str(v))
def _box(ws, S, r1, c1, r2, c2, align=None):
    """범위에 테두리 + 정렬 적용."""
    align = align or S["topleft"]
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = S["border"]
            cell.alignment = align
def _header(ws, S, cells):
    """[(row, col, value), ...] 단일 헤더 셀에 네이비 강조."""
    for (r, c, v) in cells:
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = S["hdr_fill"]
        cell.font = S["hdr_font"]
        cell.alignment = S["center"]
        cell.border = S["border"]
def _hdr_merge(ws, S, r1, c1, r2, c2, value):
    """병합 헤더 — 범위 전체 네이비/테두리, 좌상단에 값."""
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = S["hdr_fill"]
            cell.border = S["border"]
    tl = ws.cell(row=r1, column=c1, value=value)
    tl.font = S["hdr_font"]; tl.alignment = S["center"]
    return tl
def _merge_label(ws, S, r1, r2, col, value):
    """col 열의 r1~r2 병합 + 가운데 라벨 + 테두리."""
    if r2 > r1:
        ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=col).border = S["border"]
    cell = ws.cell(row=r1, column=col, value=value)
    cell.alignment = S["center"]
    return cell
def _widths(ws, mapping: dict):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w
def _build_workbook(rdf: pd.DataFrame, compare_map: dict | None = None):
    from openpyxl import Workbook
    S = _xlsx_styles()
    groups = _grouped(rdf)
    counts = _counts(rdf)
    targets = _targets(rdf)
    ranges, ds, de, footer = _ranges(groups)
    ctx = {"groups": groups, "counts": counts, "targets": targets,
           "ranges": ranges, "ds": ds, "de": de, "footer": footer,
           "compare_map": compare_map}   # 최종 보고서: 최초진단(base) {result,reason} → 3-1 시트 8컬럼

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True   # 열 때 수식 자동 재계산
    _sheet_cover(wb.active, S)
    _sheet_targets(wb.create_sheet("1. 진단 대상"), S, ctx)
    _sheet_graph(wb.create_sheet("2-1. 요약결과(그래프)"), S, ctx)
    _sheet_summary(wb.create_sheet(SUMMARY_SHEET), S, ctx)
    _sheet_detail(wb.create_sheet("3-1. 진단 결과"), S, ctx)
    return wb
def _sheet_cover(ws, S):
    ws.title = "0. 표지"
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 2})
    _widths(ws, {c: 9 for c in "BCDEFGHIJ"})
    _widths(ws, {"K": 13, "L": 30})
    for r in (3, 4, 5, 6):
        ws.row_dimensions[r].height = 22
    # 우상단 메타 박스
    meta = [("문서번호", REPORT_META["문서번호"]), ("작성자", REPORT_META["작성자"]),
            ("보안등급", REPORT_META["보안등급"]), ("Ver", REPORT_META["Ver"])]
    for i, (k, v) in enumerate(meta, start=3):
        kc = ws.cell(row=i, column=11, value=k)
        kc.font = S["Font"](bold=True); kc.border = S["border"]; kc.alignment = S["center"]
        vc = ws.cell(row=i, column=12, value=v)
        vc.border = S["border"]; vc.alignment = S["left"]
    # 제목 / 부제 / 날짜
    ws.row_dimensions[11].height = 40
    ws.row_dimensions[13].height = 30
    ws.merge_cells("B11:L11")
    t = ws.cell(row=11, column=2, value=REPORT_META["제목"])
    t.font = S["Font"](bold=True, size=28); t.alignment = S["center"]
    ws.merge_cells("B13:L13")
    st = ws.cell(row=13, column=2, value=REPORT_META["부제"])
    st.font = S["Font"](bold=True, size=16); st.alignment = S["center"]
    ws.merge_cells("B18:L18")
    dt = ws.cell(row=18, column=2, value=str(date.today()))
    dt.font = S["Font"](size=12); dt.alignment = S["center"]
def _sheet_targets(ws, S, ctx):
    targets = ctx["targets"]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 2, "B": 6, "C": 22, "D": 16, "E": 34, "F": 12, "G": 26})
    ws.row_dimensions[1].height = 26
    ws.cell(row=1, column=2,
            value="  ※ 진단 대상 리스트").font = S["Font"](bold=True)
    # 헤더 (2~3행)
    _hdr_merge(ws, S, 2, 2, 3, 2, "순번")
    _hdr_merge(ws, S, 2, 3, 2, 6, "진단 대상")
    _hdr_merge(ws, S, 2, 7, 3, 7, "비고")
    _header(ws, S, [(3, 3, "Hostname"), (3, 4, "IP Address"),
                    (3, 5, "버전정보"), (3, 6, "용도")])
    # 데이터 — 헤더 바로 아래(4행~)에 붙임
    start = 4
    for i, t in enumerate(targets):
        r = start + i
        ws.row_dimensions[r].height = 24
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value="")          # Hostname: 수기 기입용 공란
        ws.cell(row=r, column=4, value=t["ip"])
        ws.cell(row=r, column=5, value=TARGET_DEFAULTS["버전정보"])
        ws.cell(row=r, column=6, value=TARGET_DEFAULTS["용도"])
        ws.cell(row=r, column=7, value=TARGET_DEFAULTS["비고"])
    _box(ws, S, start, 2, start + len(targets) - 1, 7, align=S["center"])
    _auto_width(ws, per_col={"E": (20, 40), "G": (16, 30), "C": (14, 26)})
def _sheet_graph(ws, S, ctx):
    from openpyxl.chart import BarChart, PieChart, RadarChart, Reference
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
    from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties
    from openpyxl.styles import PatternFill
    groups, ranges = ctx["groups"], ctx["ranges"]
    ds, de = ctx["ds"], ctx["de"]
    ref = lambda a: f"'{SUMMARY_SHEET}'!{a}"   # noqa: E731
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 2, "B": 16, "C": 11})
    n = len(groups)
    STEP, SIDE = 23, 10              # 블록 간 행 간격 / 차트 정사각 한 변(cm)
    tan = PatternFill("solid", fgColor="DED9C4")   # 베이지 제목 막대

    def axis_gray(ax, size=900):
        cp = CharacterProperties(solidFill="BFBFBF", sz=size)
        ax.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    def band(row, text):
        # 제목 막대 폭을 차트 폭(F~K)에 맞춤
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
        ws.row_dimensions[row].height = 22
        for col in range(6, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = tan; cell.border = S["border"]
        c = ws.cell(row=row, column=6, value=text)
        c.font = S["Font"](bold=True, size=12); c.alignment = S["center"]

    def place(chart, top_row, span_rows=19):
        """제목 막대(F:K)와 동일한 열 범위로 차트 고정(two-cell anchor)."""
        a = TwoCellAnchor(editAs="oneCell")
        a._from = AnchorMarker(col=5, colOff=0, row=top_row - 1, rowOff=0)
        a.to = AnchorMarker(col=11, colOff=0, row=top_row - 1 + span_rows, rowOff=0)
        chart.anchor = a
        ws.add_chart(chart)

    b0, b1, b2 = 2, 2 + STEP, 2 + 2 * STEP   # 블록 시작행: 2, 25, 48

    # ── 블록1: 도메인 평균점수 (선형 방사형) ──
    band(b0, "항목별 양호 비율")
    h0 = b0 + 2
    _header(ws, S, [(h0, 2, "진단 도메인"), (h0, 3, "평균 점수")])
    for i, (d, rows, r1, r2) in enumerate(ranges):
        r = h0 + 1 + i
        ws.cell(row=r, column=2, value=d).alignment = S["center"]
        vc = ws.cell(row=r, column=3, value=f"={ref(f'H{r1}')}")
        vc.number_format = "0.0%"; vc.alignment = S["center"]
    _box(ws, S, h0 + 1, 2, h0 + n, 3, align=S["center"])
    radar = RadarChart(); radar.type = "standard"
    radar.height = radar.width = SIDE
    radar.add_data(Reference(ws, min_col=3, min_row=h0, max_row=h0 + n), titles_from_data=True)
    radar.set_categories(Reference(ws, min_col=2, min_row=h0 + 1, max_row=h0 + n))
    radar.series[0].graphicalProperties = GraphicalProperties()
    radar.series[0].graphicalProperties.line = LineProperties(solidFill="4472C4", w=28575)
    radar.x_axis.delete = False
    radar.y_axis.delete = False
    radar.y_axis.scaling.min = 0; radar.y_axis.scaling.max = 1
    radar.y_axis.majorUnit = 0.2; radar.y_axis.numFmt = "0%"
    axis_gray(radar.y_axis)
    radar.legend = None
    place(radar, b0 + 2)

    # ── 블록2: 양호/취약/N/A (원형) ──
    band(b1, "양호 / 취약 / N/A 비율")
    h1 = b1 + 2
    _header(ws, S, [(h1, 2, "상태"), (h1, 3, "개수")])
    for i, (lab, col) in enumerate([("양호", "J"), ("취약", "K"), ("N/A", "L")]):
        r = h1 + 1 + i
        ws.cell(row=r, column=2, value=lab).alignment = S["center"]
        ws.cell(row=r, column=3, value=f"=SUM({ref(f'{col}{ds}:{col}{de}')})").alignment = S["center"]
    _box(ws, S, h1 + 1, 2, h1 + 3, 3, align=S["center"])
    pie = PieChart(); pie.varyColors = True
    pie.height = pie.width = SIDE
    pie.add_data(Reference(ws, min_col=3, min_row=h1, max_row=h1 + 3), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=2, min_row=h1 + 1, max_row=h1 + 3))
    pie.legend.position = "b"
    pie.layout = Layout(manualLayout=ManualLayout(xMode="edge", yMode="edge",
                                                  x=0.22, y=0.06, w=0.56, h=0.72))
    place(pie, b1 + 2)

    # ── 블록3: 영역별 취약 수 (막대) ──
    band(b2, "영역별 취약 수")
    h2 = b2 + 2
    _header(ws, S, [(h2, 2, "영역"), (h2, 3, "취약 수")])
    for i, (d, rows, r1, r2) in enumerate(ranges):
        r = h2 + 1 + i
        ws.cell(row=r, column=2, value=d).alignment = S["center"]
        ws.cell(row=r, column=3, value=f"=SUM({ref(f'K{r1}:K{r2}')})").alignment = S["center"]
    _box(ws, S, h2 + 1, 2, h2 + n, 3, align=S["center"])
    bar = BarChart(); bar.type = "col"; bar.grouping = "clustered"; bar.varyColors = True
    bar.height = bar.width = SIDE
    bar.add_data(Reference(ws, min_col=3, min_row=h2, max_row=h2 + n), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=h2 + 1, max_row=h2 + n))
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.legend = None
    place(bar, b2 + 2)


def _sheet_summary(ws, S, ctx):
    counts, ranges = ctx["counts"], ctx["ranges"]
    ds, de, footer = ctx["ds"], ctx["de"], ctx["footer"]
    host = ctx["targets"][0]["hostname"]
    ip = ctx["targets"][0]["ip"]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 2, "B": 15, "C": 8, "D": 56, "E": 7, "F": 11, "G": 2.5,
                 "H": 11, "I": 7, "J": 7, "K": 7, "L": 7})
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B2:L2")
    tc = ws.cell(row=2, column=2,
                 value=f"{REPORT_META['제목']} 요약 진단결과 ({counts['total']}항목)")
    tc.font = S["Font"](bold=True, size=13); tc.alignment = S["center"]
    # 좌측 헤더(3~5행 세로병합)
    _hdr_merge(ws, S, 3, 2, 5, 2, "진단항목")
    _hdr_merge(ws, S, 3, 3, 5, 3, "No.")
    _hdr_merge(ws, S, 3, 4, 5, 4, "세부 진단항목")
    _hdr_merge(ws, S, 3, 5, 5, 5, "중요도")
    # 결과 컬럼(F) — 진단대상 표기 (원본: 1 / hostname / ip)
    _header(ws, S, [(3, 6, "진단결과"), (4, 6, host), (5, 6, ip)])
    # 우측 집계 헤더(5행) — 원본과 동일 레이아웃
    _header(ws, S, [(5, 8, "영역별점수"), (5, 9, "점수"),
                    (5, 10, "양호"), (5, 11, "취약"), (5, 12, "N/A")])
    # 데이터(6행~) — 모두 엑셀 수식(COUNTIF/AVERAGE) → 결과 바꾸면 자동 % 반영
    for d, rows, r1, r2 in ranges:
        for k, row in enumerate(rows):
            r = r1 + k
            res = str(row.get("결과"))
            ws.cell(row=r, column=3, value=row.get("항목코드"))
            ws.cell(row=r, column=4, value=row.get("항목"))
            ws.cell(row=r, column=5, value=row.get("중요도"))
            rc = ws.cell(row=r, column=6, value=res)
            f = _result_fill(S, res)
            if f:
                rc.fill = f
            # 점수(I): 양호=100%, 취약=0%, N/A 표기 (퍼센트)
            sc = ws.cell(row=r, column=9,
                         value=f'=IF(COUNTIF($F{r}:$F{r},"{config.R_NA}")=COUNTA($F{r}:$F{r}),'
                               f'"{config.R_NA}",$J{r}/(COUNTA($F{r}:$F{r})-$L{r}))')
            sc.number_format = "0.0%"
            ws.cell(row=r, column=10, value=f'=COUNTIF($F{r}:$F{r},"{config.R_PASS}")')
            ws.cell(row=r, column=11, value=f'=COUNTIF($F{r}:$F{r},"{config.R_VULN}")')
            ws.cell(row=r, column=12, value=f'=COUNTIF($F{r}:$F{r},"{config.R_NA}")')
        _box(ws, S, r1, 2, r2, 6, align=S["center"])
        _box(ws, S, r1, 8, r2, 12, align=S["center"])
        # 세부 진단항목(D)만 좌측정렬
        for rr in range(r1, r2 + 1):
            ws.cell(row=rr, column=4).alignment = S["leftc"]
        _merge_label(ws, S, r1, r2, 2, d)
        # 영역별점수(H) = AVERAGE(점수) — 퍼센트 (원본 방식)
        hc = _merge_label(ws, S, r1, r2, 8,
                          f'=IF(COUNTIF(I{r1}:I{r2},"{config.R_NA}")=COUNTA(I{r1}:I{r2}),'
                          f'"{config.R_NA}",AVERAGE(I{r1}:I{r2}))')
        hc.number_format = "0.0%"
    # 푸터 (수식)
    fl = _merge_label(ws, S, footer, footer + 1, 2, "점검결과")
    fl.font = S["Font"](bold=True)
    ws.merge_cells(start_row=footer, start_column=3, end_row=footer, end_column=5)
    ws.cell(row=footer, column=3, value="취약항목 개수")
    fv = ws.cell(row=footer, column=6, value=f'=COUNTIF(F{ds}:F{de},"{config.R_VULN}")')
    fv.number_format = "0"
    ws.merge_cells(start_row=footer + 1, start_column=3, end_row=footer + 1, end_column=5)
    ws.cell(row=footer + 1, column=3, value="보안 적용율 (양호항목 / 진단항목) %")
    ac = ws.cell(row=footer + 1, column=6,
                 value=f'=(COUNTIF(F{ds}:F{de},"{config.R_PASS}"))/'
                       f'(COUNTA(F{ds}:F{de})-COUNTIF(F{ds}:F{de},"{config.R_NA}"))')
    ac.number_format = "0.0%"
    _box(ws, S, footer, 2, footer + 1, 6, align=S["center"])
    ws.freeze_panes = "A6"
    _auto_width(ws, per_col={"D": (24, 60), "B": (12, 20), "E": (10, 12)})
def _sheet_detail(ws, S, ctx):
    counts, ranges = ctx["counts"], ctx["ranges"]
    compare_map = ctx.get("compare_map")   # 최종 보고서: 항목코드→{result,reason} (최초진단 base)
    is_final = compare_map is not None
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[2].height = 26
    if is_final:
        # 최종 보고서: 8컬럼 (최초/최종 결과·근거)
        _widths(ws, {"A": 2, "B": 14, "C": 9, "D": 40, "E": 60, "F": 11, "G": 40, "H": 11, "I": 40})
        last = 9
        ws.merge_cells("B2:I2")
        headers = ["진단항목", "No.", "세부 진단항목", "진단기준",
                   "최초진단결과", "최초 판단 근거", "최종진단결과", "최종 판단 근거"]
    else:
        _widths(ws, {"A": 2, "B": 14, "C": 9, "D": 44, "E": 70, "F": 10, "G": 50, "H": 50})
        last = 8
        ws.merge_cells("B2:H2")
        headers = ["진단항목", "No.", "세부 진단항목", "진단기준", "진단결과", "판단 근거", "조치 방법"]
    tc = ws.cell(row=2, column=2,
                 value=f"{REPORT_META['제목']} 상세 진단결과 ({counts['total']}항목)")
    tc.font = S["Font"](bold=True, size=13); tc.alignment = S["center"]
    # 헤더(3~5행 세로병합)
    for ci, h in enumerate(headers, start=2):
        _hdr_merge(ws, S, 3, ci, 5, ci, h)
    for d, rows, r1, r2 in ranges:
        for k, row in enumerate(rows):
            r = r1 + k
            ws.cell(row=r, column=3, value=row.get("항목코드"))
            ws.cell(row=r, column=4, value=row.get("항목"))
            ws.cell(row=r, column=5, value=_crit_lines(row.get("판단기준")))   # '|' → 줄바꿈
            if is_final:
                cm = compare_map.get(str(row.get("항목코드")), {})
                res1 = str(cm.get("result", ""))            # 최초진단(base)
                rc1 = ws.cell(row=r, column=6, value=res1)
                if _result_fill(S, res1):
                    rc1.fill = _result_fill(S, res1)
                ws.cell(row=r, column=7, value=cm.get("reason", ""))
                res2 = str(row.get("결과"))                  # 최종(이행점검 = 이 시트 기준 rdf)
                rc2 = ws.cell(row=r, column=8, value=res2)
                if _result_fill(S, res2):
                    rc2.fill = _result_fill(S, res2)
                ws.cell(row=r, column=9, value=row.get("판단근거"))
            else:
                res = str(row.get("결과"))
                rc = ws.cell(row=r, column=6, value=res)
                if _result_fill(S, res):
                    rc.fill = _result_fill(S, res)
                ws.cell(row=r, column=7, value=row.get("판단근거"))
                ws.cell(row=r, column=8, value=row.get("조치방법"))
        _box(ws, S, r1, 2, r2, last, align=S["leftc"])
        for rr in range(r1, r2 + 1):
            ws.cell(row=rr, column=3).alignment = S["center"]   # No.
            ws.cell(row=rr, column=6).alignment = S["center"]   # (최초)결과
            if is_final:
                ws.cell(row=rr, column=8).alignment = S["center"]   # 최종결과
        _merge_label(ws, S, r1, r2, 2, d)
    ws.freeze_panes = "A6"
    if is_final:
        _auto_width(ws, per_col={"D": (24, 44), "E": (45, 90), "G": (24, 50), "I": (24, 50), "B": (12, 18)})
    else:
        _auto_width(ws, per_col={"D": (24, 48), "E": (55, 100), "G": (24, 55), "H": (24, 55), "B": (12, 18)})
